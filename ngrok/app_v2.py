# -*- coding: utf-8 -*-
"""
Created on Tue Oct  1 14:24:56 2019
@author: masterp
"""

from flask import Flask, request, Response
from openpyxl import load_workbook 
import numpy as np
from ibm_watson import AssistantV2
from ibm_cloud_sdk_core.authenticators import IAMAuthenticator
import requests


############################ telegram config ########################################

API_KEY = '930278500:AAFQH3njUefoCC1iPb8wAQUQf2o2L-uGQyo'
SEND_URL = 'https://api.telegram.org/bot{token}/sendMessage'.format(token=API_KEY)
SEND_PHOTO = 'https://api.telegram.org/bot{token}/sendPhoto'.format(token=API_KEY)

############################## watson config ########################################

authenticator = IAMAuthenticator('r9GldOY6dRKQryvVfUL1MTRUcJgcDkml52jBLIAMKY3A')
service = AssistantV2(
    version='2019-10-09',
    authenticator=authenticator
)

assistant_id = '95b3bf07-4d47-4a3e-8fc6-964428aa1cf0'

session_id = service.create_session(
    assistant_id=assistant_id

).get_result()['session_id']


message_input = {
        'text':''
        }    # Watson assistant에 딕셔너리 형태로 대화 전달


############################## function definition ##################################

# 사용자 입력메세지 파싱 함수
def parse_message(message):
    
    chat_id = message['message']['chat']['id']
    msg = message['message']['text']
    
    try:
        last_name = message['message']['from']['last_name']
        first_name = message['message']['from']['first_name']
        user_name = last_name + first_name 
    except:
        user_name = "귀인"
    
    return chat_id, msg, user_name


# 사용자 inline button 클릭 callback 파싱 함수
def parse_callback(message):
    
    chat_id = message['callback_query']['message']['chat']['id']
    callback = message['callback_query']['data']

    try:
        last_name = message['callback_query']['from']['last_name']
        first_name = message['callback_query']['from']['first_name']
        user_name = last_name + first_name 
    except:
        user_name = "귀인"
    
    return chat_id, callback, user_name


# 사용자 정보를 저장하는 함수
EXCEL_FILE_NAME = './DB.xlsx'
db = load_workbook(filename=EXCEL_FILE_NAME)
user_db = db['User_DB']


def find_user_row(user_id, user_name):
    # user_exist : 사용자의 방문경험 유무(있으면 True, 없으면 False)
    # user_exist == False, is_welcome == True
    user_exist = False
    is_welcome = False


    for row in user_db.rows:
        if row[0].value == user_id:
            user_exist = True
            user_row = row[0].row
          

    if not user_exist:
        is_welcome = True
        # max_row 는 행의 최대값을 리턴해주는데,
        # 한번이라도 데이터가 적힌 행 자체를 삭제해줘야 빈칸이 안생김
        user_db[user_db.max_row + 1][0].value = user_id
        
        # 텔레그램 챗봇의 대화로 부터 사용자의 이름을 엑셀에 저장
        user_db[user_db.max_row][1].value = user_name
        
        # counter 초기화
        user_db[user_db.max_row][2].value = 0

        user_row = user_db.max_row
        
    db.save(EXCEL_FILE_NAME)
     
    return user_row, is_welcome


# 사용자가 말을 몇번 걸었는지 엑셀파일에 누적하는 함수
def user_counter(user_row):
    user_db.cell(row=user_row, column=3).value += 1
    db.save(EXCEL_FILE_NAME)
    return 0


# 처음 접속 사용자 인사말 출력 함수
def send_welcome_msg(chat_id, user_name):
    url = 'https://api.telegram.org/bot{token}/sendMessage'.format(token=API_KEY)
    welcome_msg_set = [
        "안녕하세요. 미래의 {}사장님, 꿈을 위한 지원금을 찾아드리는 청년 조(助)비서입니다. 청년수당이란? 서울에 거주하는 미취업 청년의 구직활동을 지원하기위한 수당입니다. 궁금한점이 있으시면 말씀해주세요.\n\n버튼을 이용하시면 빠르게 검색하실 수 있어요☟☟☟".format(user_name),
        "{}님의 밝은 태양 같은 미래를 지원하는 청년 조(助)비서입니다. 청년수당이란 서울에 거주하는 미취업 청년의 구직활동을 지원하기위한 수당입니다. 어떤 것이 궁금하신가요?\n\n버튼을 이용하시면 빠르게 검색하실 수 있어요☟☟☟".format(user_name),
        "{}님, 저는 서울시 청년수당에 대한 궁금증을 속 시원히 해결 해 드리는 청년 조(助)비서입니다. 청년수당이란 서울에 거주하는 미취업 청년의 구직활동을 지원하기위한 수당입니다. 궁금하신 사항에 대해서 물어봐주세요.\n\n버튼을 이용하시면 빠르게 검색하실 수 있어요☟☟☟".format(user_name),
        "{}님의 오늘보다 더 찬란한 내일을 위해 청년 조(助)비서가 달립니다. 청년수당이란 서울에 거주하는 미취업 청년의 구직활동을 지원하기위한 수당입니다. 시원한 답변이 24시간 기다리고 있습니다. 물어보세요!\n\n버튼을 이용하시면 빠르게 검색하실 수 있어요☟☟☟".format(user_name),
        "환영합니다! {}님, 주머니도 마음도 든든하게 도와드릴 청년 조(助)비서입니다. 청년수당이란 서울에 거주하는 미취업 청년의 구직활동을 지원하기위한 수당입니다. 무엇이 궁금하신가요?\n\n버튼을 이용하시면 빠르게 검색하실 수 있어요☟☟☟".format(user_name)
    ]
    params = {'chat_id': chat_id, 'text': np.random.choice(welcome_msg_set)}
    response = requests.post(url, json=params)
    
    return response


# 대화시작시 keyboard button 출력
def send_button_msg(chat_id, text):

    keyboard = {'keyboard':[[{'text': '지원혜택'},{'text': '자격요건'},{'text': '신청절차'}],
    [{'text': '선정기준'},{'text': '사용방법'},{'text': '혜택종료'}],
    [{'text': '청년 조(助)비서의 응원'},{'text': '마무으~리'}]],'one_time_keyboard' : False}

    params = {'chat_id':chat_id, 'text':text, 'reply_markup':keyboard}
    response = requests.post(SEND_URL, json=params)
    
    return response


# 대화시작시 inline keyboard button 출력
def send_inline_msg(chat_id, text):
    keyboard = {
        'inline_keyboard':[
            [{
                'text':'자가체크 바로가기',
                'url':'https://youth.seoul.go.kr/site/main/youth_allowance/self_check_step'
            }]
                ]
            }

    params = {'chat_id':chat_id, 'text':text, 'reply_markup':keyboard}
    response = requests.post(SEND_URL, json=params)
    
    return response

# 사진 전송
def send_photo(chat_id, path):
    files = {'photo':open(path, 'rb')}
    data = {'chat_id':chat_id}
    response = requests.post(SEND_PHOTO, files=files, data=data)
   
    return response

#####################################################################################


app = Flask(__name__)


@app.route('/', methods=['POST', 'GET'])
def index():
    
    if request.method == 'POST':
        message = request.get_json()

        chat_id, msg, user_name = parse_message(message)

        user_row, is_welcome = find_user_row(chat_id, user_name)
        user_counter(user_row)

        text = "(당신의 궁금증을 귀기울여 듣고 있습니다~)"
        send_button_msg(chat_id, text)

        if is_welcome:
            send_welcome_msg(chat_id, user_name)
            send_photo(chat_id, './asset/jo_profile1.jpg')
            return Response('ok', status=200)
        else:
            pass

        try:
            text = send_message(msg)        
            params = {'chat_id': chat_id, 'text': text}
            requests.post(SEND_URL, json=params)
        except:
            pass
    
        return Response('ok', status=200)

    else:
        return 'Hello World!'
        
        
        
@app.route('/send_message/<message>')
def send_message(message):
    '''
    왓슨으로 message를 보내줍니다
    :param message: telegram으로부터 받은 사용자의 input
    :return:
    '''
    message_input['text'] = message

    #왓슨으로 message를 보내고 결과까지 받기
    response = service.message(
        assistant_id = assistant_id,
        session_id = session_id,
        input = message_input
    ).get_result()


    # intent 가져오기
    if response['output']['intents']:
        print(response['output']['intents'][0]['intent'])

    # 왓슨에서 return한 response가져오기
    if response['output']['generic']:
        if response['output']['generic'][0]['response_type'] == 'text':
            return(response['output']['generic'][0]['text'])




@app.route('/about')
def about():
  return 'About page'


if __name__ == '__main__':
    app.run(port = 5000)