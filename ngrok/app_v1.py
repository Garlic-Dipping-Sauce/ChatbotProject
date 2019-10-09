# -*- coding: utf-8 -*-
"""
Created on Tue Oct  1 14:24:56 2019
@author: masterp
"""

from flask import Flask, request, Response
from openpyxl import load_workbook 
import numpy as np
from ibm_watson import AssistantV2
from ibm_cloud_sdk_core.authenticators import IAMAuthenticator
import requests


############################ telegram config ########################################

API_KEY = '930278500:AAFQH3njUefoCC1iPb8wAQUQf2o2L-uGQyo'
SEND_URL = 'https://api.telegram.org/bot{token}/sendMessage'.format(token=API_KEY)
SEND_PHOTO = 'https://api.telegram.org/bot{token}/sendPhoto'.format(token=API_KEY)

############################## watson config ########################################

authenticator = IAMAuthenticator('yEHYkbVBGBfoOukgGRky30SZyqnudr4YHfkOfHJGHgIT')
service = AssistantV2(
    version='2019-10-07',
    authenticator=authenticator
)

assistant_id = '93208712-5231-4e4f-ab65-c1da90a8b245'

session_id = service.create_session(
    assistant_id=assistant_id

).get_result()['session_id']


message_input = {
        'text':''
        }    # Watson assistant에 딕셔너리 형태로 대화 전달


############################## function definition ##################################

# 사용자 입력메세지 파싱 함수
def parse_message(message):
    
    chat_id = message['message']['chat']['id']
    msg = message['message']['text']
    
    try:
        last_name = message['message']['from']['last_name']
        first_name = message['message']['from']['first_name']
        user_name = last_name + first_name 
    except:
        user_name = "귀인"
    
    return chat_id, msg, user_name


# 사용자 inline button 클릭 callback 파싱 함수
def parse_callback(message):
    
    chat_id = message['callback_query']['message']['chat']['id']
    callback = message['callback_query']['data']

    try:
        last_name = message['callback_query']['from']['last_name']
        first_name = message['callback_query']['from']['first_name']
        user_name = last_name + first_name 
    except:
        user_name = "귀인"
    
    return chat_id, callback, user_name


# 사용자 정보를 저장하는 함수
EXCEL_FILE_NAME = './DB.xlsx'
db = load_workbook(filename=EXCEL_FILE_NAME)
user_db = db['User_DB']


def find_user_row(user_id, user_name):
    # user_exist : 사용자의 방문경험 유무(있으면 True, 없으면 False)
    # user_exist == False, is_welcome == True
    user_exist = False
    is_welcome = False


    for row in user_db.rows:
        if row[0].value == user_id:
            user_exist = True
            user_row = row[0].row
          

    if not user_exist:
        is_welcome = True
        # max_row 는 행의 최대값을 리턴해주는데,
        # 한번이라도 데이터가 적힌 행 자체를 삭제해줘야 빈칸이 안생김
        user_db[user_db.max_row + 1][0].value = user_id
        
        # 텔레그램 챗봇의 대화로 부터 사용자의 이름을 엑셀에 저장
        user_db[user_db.max_row][1].value = user_name
        
        # counter 초기화
        user_db[user_db.max_row][2].value = 0

        user_row = user_db.max_row
        
    db.save(EXCEL_FILE_NAME)
     
    return user_row, is_welcome


# 사용자가 말을 몇번 걸었는지 엑셀파일에 누적하는 함수
def user_counter(user_row):
    user_db.cell(row=user_row, column=3).value += 1
    db.save(EXCEL_FILE_NAME)
    return 0


# 처음 접속 사용자 인사말 출력 함수
def send_welcome_msg(chat_id, user_name):
    url = 'https://api.telegram.org/bot{token}/sendMessage'.format(token=API_KEY)
    welcome_msg_set = [
        "안녕하세요. 미래의 건물주 {}님의 현재 지갑을 채워줄 청년 조비서입니다. 청년수당에 대해선 뭐든 다 알고 있으니 편안하게 물어봐주세요.".format(user_name),
        "안녕하세요. {}님이 미쳐 챙기지 못한 숨겨진 지원금에 대해서 자세하고 명쾌하게 알려드릴 청년 조비서입니다. 어떤 질문이라도 답 해드릴게요.".format(user_name),
        "안녕하세요. 눈부신 태양처럼 밝은 {}님의 미래를 위해 준비된 청년수당 전문가 청년 조비서입니다. 궁금하신 사항이 있으시면 질문해주세요.".format(user_name)
    ]
    params = {'chat_id': chat_id, 'text': np.random.choice(welcome_msg_set)}
    requests.post(url, json=params)

# 대화시작시 inline keyboard button 출력
def send_inline_msg(chat_id, text):
    keyboard = {
        'inline_keyboard':[
            [{
                'text':'지원혜택',
                'callback_data':'지원혜택'
            },
            {
                'text':'자격요건',
                'callback_data':'자격요건'
            },
            {
                'text':'신청방법',
                'callback_data':'신청방법'
            }],
            [{
                'text':'선정기준',
                'callback_data':'선정기준'
            },
            {
                'text':'사용방법',
                'callback_data':'사용방법'
            },
            {
                'text':'혜택종료',
                'callback_data':'혜택종료'
            }],
            [{
                'text':'자가체크 바로가기',
                'url':'https://youth.seoul.go.kr/site/main/youth_allowance/self_check_step'
            },
            {
                'text':'마무의~리',
                'callback_data':'끝인사'
            }]
                ]
            }

    params = {'chat_id':chat_id, 'text':text, 'reply_markup':keyboard}
    response = requests.post(SEND_URL, json=params)
    
    return response

# 사진 전송
def send_photo(chat_id, path):
     
    params = {'chat_id':chat_id, 'photo':open('path', 'rb')}
    response = requests.post(SEND_PHOTO, json=params)
    
    return response

#####################################################################################


app = Flask(__name__)


@app.route('/', methods=['POST', 'GET'])
def index():
    
    if request.method == 'POST':
        message = request.get_json()
    
        if 'callback_query' in message:
            chat_id, msg, user_name = parse_callback(message)
            print(chat_id, msg, user_name)        
        else:
            chat_id, msg, user_name = parse_message(message)


        user_row, is_welcome = find_user_row(chat_id, user_name)

        user_counter(user_row)

        if is_welcome:
            send_welcome_msg(chat_id, user_name)
            return Response('ok', status=200)
        
        text = "무엇을 물어야할지 모르시면 질문을 눌러보세요"
        send_inline_msg(chat_id, text)

        text = send_message(msg)
        
        params = {'chat_id': chat_id, 'text': text}
        requests.post(SEND_URL, json=params)
        
        return Response('ok', status=200)

    else:
        return 'Hello World!'
        
        
        
@app.route('/send_message/<message>')
def send_message(message):
    '''
    왓슨으로 message를 보내줍니다
    :param message: telegram으로부터 받은 사용자의 input
    :return:
    '''
    message_input['text'] = message

    #왓슨으로 message를 보내고 결과까지 받기
    response = service.message(
        assistant_id = assistant_id,
        session_id = session_id,
        input = message_input
    ).get_result()


    # intent 가져오기
    if response['output']['intents']:
        print(response['output']['intents'][0]['intent'])

    # 왓슨에서 return한 response가져오기
    if response['output']['generic']:
        if response['output']['generic'][0]['response_type'] == 'text':
            return(response['output']['generic'][0]['text'])




@app.route('/about')
def about():
  return 'About page'


if __name__ == '__main__':
    app.run(port = 5000)